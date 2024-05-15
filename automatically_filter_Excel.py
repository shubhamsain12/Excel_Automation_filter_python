import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
from openpyxl import Workbook

# URL of the login page
login_url = "https://power.bigrock247.com/admin/login"
username = "jeetudelhi"
password = "Jeetu@1234#"

# Custom "From" and "To" dates
from_date = "2024-05-11"
to_date = "2024-05-14"

# Create a new instance of the Chrome driver
driver = webdriver.Chrome()

try:
    # Load the login page
    driver.get(login_url)

    # Find the username and password input fields and enter the credentials
    username_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="user_name"]')))
    password_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="passwordGroup"]/div/input')))

    username_input.send_keys(username)
    password_input.send_keys(password)

    # Find and click the login button
    login_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[2]/div/form/button')))
    login_button.click()

    print("Successfully logged in.")

    # Find and click the "Account" button on the admin page
    account_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/nav/div/div[2]/ul/li/a')))
    account_button.click()

    # Now, wait for the drop-down menu to appear
    dropdown_menu = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/nav/div/div[2]/ul/li/ul')))

    # Click on the "My Bets" option in the dropdown menu
    my_bets_option = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/nav/div/div[2]/ul/li/ul/li[4]')))
    my_bets_option.click()
    
    # Wait for the page to load
    time.sleep(5)

    # Click on the "From Date" element and enter the custom "From" date using JavaScript
    driver.execute_script(f"document.getElementById('from_date').value = '{from_date}';")

    # Click on the "To Date" element and enter the custom "To" date using JavaScript
    driver.execute_script(f"document.getElementById('to_date').value = '{to_date}';")

    # Wait for 5 seconds for the date range to update
    time.sleep(5)

    # Find and click the "Completed" button
    completed_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btn_completed"]')))
    completed_button.click()

    # Wait for the "Running Matches" table to load
    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="runningMatches"]')))

    # Find the "Show entries" dropdown menu and click on it
    show_entries_dropdown = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="runningMatches_length"]/label/select')))
    show_entries_dropdown.click()

    # Find and click the "100" option from the dropdown menu
    option_100 = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="runningMatches_length"]/label/select/option[4]')))
    option_100.click()

    # Wait for 3 seconds for the table to update
    time.sleep(3)

    # Initialize an empty DataFrame to store the scraped data
    all_data = pd.DataFrame()

    while True:
        try:
            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(driver.page_source, "html.parser")

            # Extract data from the table
            data = []
            for row in soup.find("table", id="runningMatches").find_all("tr"):
                cells = row.find_all("td")
                if len(cells) > 0:
                    description = cells[0].text.strip()
                    market = cells[1].text.strip()
                    odd = cells[2].text.strip()
                    stack = cells[3].text.strip()
                    profit_loss = cells[4].text.strip()
                    status = cells[5].text.strip()
                    ip_address = cells[6].text.strip()
                    data.append([description, market, odd, stack, profit_loss, status, ip_address])

            # Create a DataFrame with the extracted data
            columns = ['Description', 'Market', 'Odd', 'Stack', 'Profit Loss', 'Status', 'IP Address']
            df = pd.DataFrame(data, columns=columns)

            # Append the data to the main DataFrame
            all_data = pd.concat([all_data, df], ignore_index=True)

            # Save DataFrame to an Excel file
            excel_file = "NEW.xlsx"
            all_data.to_excel(excel_file, index=False)

            print(f"Data saved to {excel_file} successfully.")

            # Check if there is a "Next" button and if it's not disabled
            next_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="runningMatches_next"]')))
            if "disabled" in next_button.get_attribute("class"):
                print("No more pages to scrape.")
                break  # Exit the loop if the "Next" button is disabled

            next_button.click()

            # Wait for the data on the next page to load
            WebDriverWait(driver, 30).until(EC.staleness_of(next_button))  # Wait for the next button to become stale

            # Wait for 5 seconds for the new data to load
            time.sleep(5)

        except (TimeoutException, NoSuchElementException) as e:
            print(f"Error occurred: {e}")
            break

    # Load the workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"

    for r_idx, row in enumerate(all_data.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    print("All data added to 'All Data' sheet successfully.")

    # Define a function to create and save filtered sheets
    def save_filtered_sheet(wb, all_data, filter_value, sheet_name):
        filtered_data = all_data[all_data['Market'].str.contains(filter_value, case=False, na=False)]
        if not filtered_data.empty:
            new_sheet = wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(filtered_data.values, 1):
                for c_idx, value in enumerate(row, 1):
                    new_sheet.cell(row=r_idx, column=c_idx, value=value)
            print(f"Data filtered and saved to {sheet_name} sheet successfully.")
    
    # Apply the filtering and save to new sheets
    save_filtered_sheet(wb, all_data, "Fancy", "Fancy")
    save_filtered_sheet(wb, all_data, "Match Odds", "Match Odds")
    save_filtered_sheet(wb, all_data, "Book Maker", "Book Maker")

    # Save the workbook with the new sheets
    wb.save(excel_file)
    print(f"Filtered data saved to new sheets in {excel_file} successfully.")

finally:
    # Close the browser window
    driver.quit()
